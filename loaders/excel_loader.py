from pathlib import Path
from typing import Dict, List, Union
import pandas as pd
from openpyxl import load_workbook
from langchain_core.documents import Document

from .tabular_loader import TabularLoader, TabularMetadata

class ExcelLoader(TabularLoader):
    """Loader for Excel files (XLSX and XLS formats)."""
    
    def __init__(self, file_path: Union[str, Path]):
        """Initialize the Excel loader.
        
        Args:
            file_path: Path to the Excel file
        """
        self.file_path = Path(file_path)
        if not self.file_path.exists():
            raise FileNotFoundError(f"Excel file not found: {self.file_path}")
        
        # Validate file extension
        if self.file_path.suffix.lower() not in ['.xlsx', '.xls']:
            raise ValueError(f"Unsupported file format: {self.file_path.suffix}. Must be .xlsx or .xls")
        
        # Cache metadata
        self._metadata = None
    
    def load(self) -> List[Document]:
        """Load data from the Excel file and convert to LangChain documents.
        
        Returns:
            List of LangChain Document objects, one per sheet.
        """
        metadata = self.get_metadata()
        docs = []
        
        # Read all sheets
        for sheet_name in metadata.sheet_names:
            try:
                df = pd.read_excel(
                    self.file_path,
                    sheet_name=sheet_name,
                    engine='openpyxl' if self.file_path.suffix.lower() == '.xlsx' else 'xlrd'
                )
                
                # Skip empty sheets
                if len(df) == 0:
                    continue
                
                # Create one document per sheet
                docs.append(Document(
                    page_content=df.to_string(index=False),
                    metadata={
                        "source": str(self.file_path),
                        "file_type": metadata.source_type,
                        "sheet_name": sheet_name,
                        "total_rows": len(df),
                        "total_columns": len(df.columns),
                        "columns": list(df.columns)
                    }
                ))
                
            except Exception as e:
                # Log error but continue processing other sheets
                print(f"Error reading sheet '{sheet_name}': {str(e)}")
                continue
        
        if not docs:
            raise ValueError(f"No valid data found in Excel file: {self.file_path}")
        
        return docs
    
    def get_metadata(self) -> TabularMetadata:
        """Extract metadata from the Excel file.
        
        Returns:
            TabularMetadata object containing file information.
        """
        if self._metadata is not None:
            return self._metadata
        
        # Get sheet names
        if self.file_path.suffix.lower() == '.xlsx':
            wb = load_workbook(self.file_path, read_only=True)
            sheet_names = wb.sheetnames
            wb.close()
        else:
            # For .xls files, use pandas to get sheet names
            sheet_names = pd.ExcelFile(self.file_path).sheet_names
        
        # Read first row of each sheet to get column info
        total_rows = 0
        all_columns = set()
        column_types = {}
        
        for sheet_name in sheet_names:
            try:
                # Read sheet with pandas
                df = pd.read_excel(
                    self.file_path,
                    sheet_name=sheet_name,
                    engine='openpyxl' if self.file_path.suffix.lower() == '.xlsx' else 'xlrd'
                )
                
                # Update totals
                total_rows += len(df)
                all_columns.update(df.columns)
                
                # Update column types
                sheet_types = self._get_column_types(df)
                column_types.update(sheet_types)
                
            except Exception as e:
                print(f"Error reading metadata from sheet '{sheet_name}': {str(e)}")
                continue
        
        self._metadata = TabularMetadata(
            source_type='xlsx' if self.file_path.suffix.lower() == '.xlsx' else 'xls',
            sheet_names=sheet_names,
            total_rows=total_rows,
            total_columns=len(all_columns),
            column_types=column_types,
            file_path=str(self.file_path),
            additional_metadata={
                'file_size': self.file_path.stat().st_size,
                'last_modified': self.file_path.stat().st_mtime
            }
        )
        
        return self._metadata

if __name__ == "__main__":
    import sys
    if len(sys.argv) > 1:
        file_path = sys.argv[1]
        loader = ExcelLoader(file_path)
        docs = loader.load()
        print(f"Loaded {len(docs)} documents")
        print(docs)
        if docs:
            print("\nFirst document preview:")
            doc = docs[0]
            print(f"Metadata: {doc.metadata}")
            print(f"Content preview: {doc.page_content[:200]}...") 